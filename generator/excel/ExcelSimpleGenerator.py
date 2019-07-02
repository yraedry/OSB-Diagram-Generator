from openpyxl import Workbook
from openpyxl import load_workbook
import utils.DirectoriesOperations
import utils.FilesOperations
import os


def excel_simple_generator(service_name):
    service_object = utils.DirectoriesOperations.navigate_path(service_name, type="VS")
    check_object = False
    dest_filename = 'C:\\Users\\anunez\\Documents\\Generador_umls\\excels\\CatalogoDeServicios.xlsx'
    try:
        wb = load_workbook(filename='C:\\Users\\anunez\\Documents\\Generador_umls\\excels\\CatalogoDeServicios.xlsx')
    except:
        # creo una nueva hoja
        wb = Workbook()
        wb.remove(wb['Sheet'])
        print('fichero no encontrado')

    # Nombre de la hoja excel
    ws1 = wb.create_sheet(title=service_name)
    # Write some data headers.
    ws1['A1'] = 'Service_Name'
    ws1['B1'] = 'Operation_Name'
    ws1['C1'] = 'Business_Name'
    count_proxy_colum = 2
    count_business_colum = 2
    count_operation_colum = 2

    while check_object is False:
        for excel_object in service_object:
            print(excel_object)
            for excel_object_name in excel_object:
                if '.proxy' in excel_object_name:
                    search_colum = 'A' + str(count_proxy_colum)
                    ws1[search_colum] = excel_object_name
                    count_proxy_colum = count_proxy_colum + 1
                if '.bix' in excel_object_name:
                    search_colum = 'C' + str(count_business_colum)
                    ws1[search_colum] = excel_object_name
                    count_business_colum = count_business_colum + 1
                if not excel_object_name.endswith(('.proxy', '.pipeline', '.bix')):
                    search_colum = 'B' + str(count_operation_colum)
                    ws1[search_colum] = excel_object_name
                    count_operation_colum = count_operation_colum + 1
        wb.save(filename=dest_filename)
        wb.close()
        check_object = True


if __name__ == "__main__":
    os.chdir('C:\\proyectos\\commons\\python\\CrossReferences-DiagramGenerator\\github')
    file = open('services.txt', "r")
    fl = file.readlines()
    for x in fl:
        excel_simple_generator(x.rstrip('\n'))
